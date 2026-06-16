"""
Contract Review Agent — Zalopay FP&A
======================================
Trích xuất & review hợp đồng Word + so sánh với email alignment.

Yêu cầu:
    pip install python-docx openpyxl anthropic

Sử dụng:
    # Chỉ hợp đồng:
    python contract_extractor_v2.py --contracts hop_dong.docx phu_luc.docx

    # Hợp đồng + email để so sánh:
    python contract_extractor_v2.py --contracts hop_dong.docx --emails aligned.eml

    # Toàn bộ folder:
    python contract_extractor_v2.py --folder ./ho_so/

    # Chỉ định output:
    python contract_extractor_v2.py --contracts hop_dong.docx --output review.xlsx
"""

import sys, os, json, argparse, email, re
from pathlib import Path
from datetime import datetime
from email import policy as email_policy

try:
    from docx import Document
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
except ImportError:
    sys.exit("Thiếu thư viện: pip install python-docx")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Thiếu thư viện: pip install openpyxl")

try:
    import anthropic
except ImportError:
    sys.exit("Thiếu thư viện: pip install anthropic")


# ═══════════════════════════════════════════════
# PHẦN 1: ĐỌC FILE
# ═══════════════════════════════════════════════

def read_docx(path: str) -> str:
    """Đọc toàn bộ text từ .docx, bao gồm bảng."""
    doc = Document(path)
    parts = []
    for block in doc.element.body:
        tag = block.tag.split("}")[-1]
        if tag == "p":
            text = "".join(
                node.text or "" for node in block.iter() if node.tag == qn("w:t")
            )
            if text.strip():
                parts.append(text)
        elif tag == "tbl":
            tbl = DocxTable(block, doc)
            for row in tbl.rows:
                row_text = " | ".join(cell.text.strip() for cell in row.cells)
                if row_text.strip(" |"):
                    parts.append(f"[TABLE] {row_text}")
    return "\n".join(parts)


def read_eml(path: str) -> str:
    """Đọc nội dung email .eml — lấy plain text hoặc fallback HTML stripped."""
    with open(path, "rb") as f:
        msg = email.message_from_bytes(f.read(), policy=email_policy.default)

    subject = msg.get("Subject", "")
    sender  = msg.get("From", "")
    date    = msg.get("Date", "")
    to      = msg.get("To", "")

    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                body += part.get_content() or ""
            elif ct == "text/html" and not body:
                html = part.get_content() or ""
                body += re.sub(r"<[^>]+>", " ", html)
                body = re.sub(r"\s+", " ", body)
    else:
        body = msg.get_content() or ""

    return (
        f"[EMAIL]\nTừ: {sender}\nGửi đến: {to}\nNgày: {date}\n"
        f"Tiêu đề: {subject}\n\nNội dung:\n{body}"
    )


def collect_files(contracts: list, emails: list, folder: str | None) -> tuple[dict, dict]:
    """Trả về (contract_texts, email_texts)."""
    contract_texts = {}
    email_texts = {}

    if folder:
        p = Path(folder)
        contracts = contracts + [str(f) for f in p.glob("*.docx")]
        emails    = emails    + [str(f) for f in p.glob("*.eml")]

    for fp in contracts:
        name = Path(fp).name
        try:
            contract_texts[name] = read_docx(fp)
            print(f"  ✓ HĐ/PL: {name}")
        except Exception as e:
            print(f"  ✗ Lỗi {name}: {e}")

    for fp in emails:
        name = Path(fp).name
        try:
            email_texts[name] = read_eml(fp)
            print(f"  ✓ Email: {name}")
        except Exception as e:
            print(f"  ✗ Lỗi {name}: {e}")

    return contract_texts, email_texts


# ═══════════════════════════════════════════════
# PHẦN 2: CLAUDE EXTRACTION
# ═══════════════════════════════════════════════

SYSTEM_PROMPT = """Bạn là chuyên gia phân tích hợp đồng cho bộ phận FP&A của Zalopay.
Nhiệm vụ: Đọc nội dung hợp đồng và phụ lục, trích xuất chính xác các thông tin theo cấu trúc JSON.
Quy tắc bắt buộc:
- Trả về JSON hợp lệ, KHÔNG thêm giải thích ngoài JSON.
- Rephrase lại nội dung concisely nhưng KHÔNG thay đổi ý nghĩa.
- Nếu thông tin không có: dùng null. Nếu thông tin không rõ ràng: ghi "Cần xác nhận thêm".
- Đối với phí và số tiền: ghi đầy đủ đơn vị tiền tệ và tỷ lệ %.
- Với điều khoản pháp lý: trích dẫn ngắn gọn, đủ ý."""

EXTRACTION_PROMPT = """Phân tích hợp đồng/phụ lục dưới đây. Trả về JSON theo cấu trúc này:

{{
  "loai_hop_dong": "string từ tiêu đề",

  "cac_ben": [
    {{
      "ten_ben": "Bên A / Bên B / tên gọi trong HĐ",
      "ten_cong_ty": "tên đầy đủ",
      "nguoi_dai_dien": "họ tên",
      "chuc_vu": "chức vụ",
      "dia_chi": "địa chỉ nếu có",
      "ma_so_thue": "MST nếu có"
    }}
  ],

  "thoi_han_hop_dong": {{
    "ngay_ky": "DD/MM/YYYY hoặc null",
    "ngay_hieu_luc": "DD/MM/YYYY hoặc null",
    "thoi_gian_hieu_luc": "VD: 12 tháng kể từ ngày hiệu lực",
    "ngay_het_han": "DD/MM/YYYY hoặc null",
    "dieu_kien_gia_han": "mô tả concise về gia hạn tự động/thủ công, thời hạn thông báo v.v."
  }},

  "dich_vu_hop_tac": {{
    "mo_ta_chung": "mô tả ngành nghề và phạm vi hợp tác",
    "doi_tuong_tich_hop": "đối tượng/sản phẩm tích hợp",
    "kenh_thanh_toan": ["Zalopay App", "Gateway", "VietQR", "..."],
    "nguon_tien": ["Số dư ví", "Tài khoản ngân hàng", "Thẻ nội địa", "Thẻ quốc tế", "..."]
  }},

  "commercial_terms": {{
    "tong_gia_tri_hop_dong": "số tiền + đơn vị hoặc null",

    "phi_dich_vu": [
      {{
        "loai_phi": "tên loại phí",
        "kenh_thanh_toan": "kênh áp dụng",
        "nguon_tien": "nguồn tiền áp dụng",
        "muc_phi": "con số/tỷ lệ cụ thể",
        "dieu_kien": "điều kiện áp dụng ngắn gọn"
      }}
    ],

    "ngan_sach_khuyen_mai": {{
      "tong_ngan_sach": "số tiền hoặc null",
      "the_le_dieu_kien": "mô tả thể lệ, điều kiện áp dụng hoặc null"
    }},

    "lai_va_phat": {{
      "lai_tra_cham": "mức lãi và cách tính",
      "phat_vi_pham": "mức phạt vi phạm",
      "boi_thuong_thiet_hai": "quy định bồi thường"
    }},

    "payment_term": {{
      "co_che_thanh_toan": "mô tả cơ chế: cấn trừ D+1 / đối tác TT sau X ngày từ HĐ đối soát / v.v.",
      "tam_ung_thanh_toan_truoc": "số tiền tạm ứng/TT trước nếu có, hoặc null",
      "cong_no_thanh_toan": "X ngày thường / X ngày làm việc",
      "ho_so_thanh_toan": ["hóa đơn VAT", "biên bản đối soát", "biên bản nghiệm thu", "..."],
      "alert_ho_so": "CẢNH BÁO nếu thiếu hóa đơn/biên bản đối soát/nghiệm thu, hoặc null"
    }},

    "reconciliation_term": {{
      "thoi_gian_bat_dau_doi_soat": "mô tả",
      "thoi_gian_gui_doi_soat": "mô tả",
      "thoi_gian_phan_hoi": "mô tả",
      "xu_ly_chenh_lech": "quy trình xử lý chênh lệch",
      "zalopay_xuat_hoa_don": "X ngày kể từ khi 2 bên xác nhận đối soát, hoặc null"
    }}
  }},

  "truong_con_thieu": ["danh sách các trường quan trọng không tìm thấy trong HĐ"]
}}

=== NỘI DUNG TÀI LIỆU ===
{content}
"""

COMPARISON_PROMPT = """So sánh nội dung đã aligned trong EMAIL với nội dung trong HỢP ĐỒNG.
Trả về JSON:

{{
  "ket_qua_so_sanh": [
    {{
      "diem_so_sanh": "tên điểm so sánh (VD: Mức phí dịch vụ)",
      "trong_email": "nội dung aligned trong email",
      "trong_hop_dong": "nội dung trong HĐ",
      "trang_thai": "KHỚP / KHÁC BIỆT / CHỈ TRONG EMAIL / CHỈ TRONG HĐ",
      "ghi_chu": "giải thích ngắn nếu có sai lệch"
    }}
  ],
  "tong_ket": "nhận xét tổng quan về mức độ khớp giữa email và HĐ"
}}

=== NỘI DUNG EMAIL ===
{email_content}

=== NỘI DUNG HỢP ĐỒNG (tóm tắt) ===
{contract_summary}
"""


def call_claude(prompt: str, system: str, api_key: str, model="claude-opus-4-8") -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model=model,
        max_tokens=6000,
        system=system,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = msg.content[0].text.strip()
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def extract_contract_info(contract_texts: dict, api_key: str) -> dict:
    full = ""
    for name, text in contract_texts.items():
        full += f"\n\n{'='*60}\nTÀI LIỆU: {name}\n{'='*60}\n{text}"
    prompt = EXTRACTION_PROMPT.format(content=full[:130000])
    print("  → Trích xuất thông tin HĐ qua Claude API...")
    return call_claude(prompt, SYSTEM_PROMPT, api_key)


def compare_email_contract(email_texts: dict, contract_data: dict, api_key: str) -> dict:
    email_full = "\n\n".join(email_texts.values())[:40000]
    contract_summary = json.dumps(contract_data, ensure_ascii=False, indent=2)[:30000]
    prompt = COMPARISON_PROMPT.format(
        email_content=email_full,
        contract_summary=contract_summary
    )
    print("  → So sánh email vs hợp đồng qua Claude API...")
    return call_claude(prompt, SYSTEM_PROMPT, api_key)


# ═══════════════════════════════════════════════
# PHẦN 3: EXCEL OUTPUT
# ═══════════════════════════════════════════════

# Màu sắc
C_DARK_BLUE   = "1F4E79"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_VERY_LIGHT  = "EBF3FB"
C_WHITE       = "FFFFFF"
C_RED_HEAD    = "C00000"
C_RED_LIGHT   = "FDECEA"
C_YELLOW      = "FFF2CC"
C_GREEN_LIGHT = "E2EFDA"
C_ORANGE      = "F4B942"
C_GRAY        = "F2F2F2"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="B0C4D8")
    return Border(left=s, right=s, top=s, bottom=s)

def _red_border():
    s = Side(style="medium", color="C00000")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header_main(cell, text, bg=C_DARK_BLUE, size=13):
    cell.value = text
    cell.font = _font(bold=True, size=size, color=C_WHITE)
    cell.fill = _fill(bg)
    cell.alignment = _align(h="center", v="center")

def style_section(cell, text, bg=C_MID_BLUE):
    cell.value = text
    cell.font = _font(bold=True, size=10, color=C_WHITE)
    cell.fill = _fill(bg)
    cell.alignment = _align(h="left", v="center")

def style_sub_section(cell, text):
    cell.value = text
    cell.font = _font(bold=True, size=10, color=C_DARK_BLUE)
    cell.fill = _fill(C_LIGHT_BLUE)
    cell.alignment = _align(h="left", v="center")

def style_label(cell, text):
    cell.value = text
    cell.font = _font(bold=True, size=10)
    cell.fill = _fill(C_VERY_LIGHT)
    cell.alignment = _align()
    cell.border = _border()

def style_value(cell, text, alert=False):
    cell.value = text or ""
    cell.font = _font(size=10, color="C00000" if alert else "000000")
    cell.fill = _fill(C_RED_LIGHT if alert else C_WHITE)
    cell.alignment = _align()
    cell.border = _red_border() if alert else _border()

def style_table_header(cell, text):
    cell.value = text
    cell.font = _font(bold=True, size=9, color=C_WHITE)
    cell.fill = _fill(C_MID_BLUE)
    cell.alignment = _align(h="center", v="center")
    cell.border = _border()

def style_table_cell(cell, text, alt=False, red=False):
    cell.value = text or ""
    cell.font = _font(size=9, color="C00000" if red else "000000")
    cell.fill = _fill(C_RED_LIGHT if red else (C_GRAY if alt else C_WHITE))
    cell.alignment = _align()
    cell.border = _border()

def write_section_row(ws, row, label, value, alert=False, row_height=None):
    lc = ws.cell(row=row, column=1)
    vc = ws.cell(row=row, column=2)
    style_label(lc, label)
    style_value(vc, value, alert=alert)
    if row_height:
        ws.row_dimensions[row].height = row_height
    else:
        ws.row_dimensions[row].height = 40 if value and len(str(value)) > 100 else 20

def merge_section_header(ws, row, text, cols=2, bg=C_MID_BLUE):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1)
    style_section(c, text, bg=bg)
    ws.row_dimensions[row].height = 22
    return row + 1

def merge_sub_header(ws, row, text, cols=2):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1)
    style_sub_section(c, text)
    ws.row_dimensions[row].height = 20
    return row + 1


# ─── Sheet 1: Tóm tắt HĐ ───

def write_summary_sheet(ws, data: dict, source_files: list):
    ws.title = "📋 Tóm tắt HĐ"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 72
    row = 1

    # Tiêu đề
    ws.merge_cells(f"A{row}:B{row}")
    style_header_main(ws.cell(row, 1), "BẢNG TRÍCH XUẤT THÔNG TIN HỢP ĐỒNG — ZALOPAY FP&A")
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1,
        value=f"📅 Ngày xử lý: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
              f"📄 Nguồn: {', '.join(Path(f).name for f in source_files)}")
    c.font = _font(italic=True, size=9, color="666666")
    c.alignment = _align(h="left", v="center", wrap=False)
    ws.row_dimensions[row].height = 18
    row += 1

    # Alert thiếu thông tin
    missing = data.get("truong_con_thieu") or []
    if missing:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, value=f"⚠️  THÔNG TIN CÒN THIẾU: {' | '.join(missing)}")
        c.font = _font(bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_RED_HEAD)
        c.alignment = _align(h="left", v="center")
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ── 1. Loại HĐ ──
    row = merge_section_header(ws, row, "1.  LOẠI HỢP ĐỒNG")
    write_section_row(ws, row, "Loại hợp đồng", data.get("loai_hop_dong"),
                      alert=not data.get("loai_hop_dong"))
    row += 2

    # ── 2. Các Bên ──
    row = merge_section_header(ws, row, "2.  CÁC BÊN THAM GIA")
    for p in (data.get("cac_ben") or []):
        style_label(ws.cell(row, 1), f"Bên: {p.get('ten_ben','')}")
        style_value(ws.cell(row, 2), p.get("ten_cong_ty"))
        ws.row_dimensions[row].height = 18; row += 1
        write_section_row(ws, row, "  Người đại diện",
            f"{p.get('nguoi_dai_dien','')} — {p.get('chuc_vu','')}")
        row += 1
        if p.get("ma_so_thue"):
            write_section_row(ws, row, "  MST", p.get("ma_so_thue"))
            row += 1
    row += 1

    # ── 3. Thời hạn ──
    row = merge_section_header(ws, row, "3.  THỜI HẠN HỢP ĐỒNG")
    t = data.get("thoi_han_hop_dong") or {}
    for label, key in [
        ("Ngày ký", "ngay_ky"),
        ("Ngày hiệu lực", "ngay_hieu_luc"),
        ("Thời gian hiệu lực", "thoi_gian_hieu_luc"),
        ("Ngày hết hạn", "ngay_het_han"),
        ("Điều kiện gia hạn", "dieu_kien_gia_han"),
    ]:
        write_section_row(ws, row, label, t.get(key), alert=not t.get(key))
        row += 1
    row += 1

    # ── 4. Dịch vụ hợp tác ──
    row = merge_section_header(ws, row, "4.  DỊCH VỤ HỢP TÁC")
    dv = data.get("dich_vu_hop_tac") or {}
    write_section_row(ws, row, "Mô tả chung", dv.get("mo_ta_chung"), row_height=50); row += 1
    write_section_row(ws, row, "Đối tượng tích hợp", dv.get("doi_tuong_tich_hop")); row += 1
    write_section_row(ws, row, "Kênh thanh toán",
        ", ".join(dv.get("kenh_thanh_toan") or []) or None,
        alert=not dv.get("kenh_thanh_toan")); row += 1
    write_section_row(ws, row, "Nguồn tiền",
        ", ".join(dv.get("nguon_tien") or []) or None,
        alert=not dv.get("nguon_tien")); row += 1
    row += 1

    # ── 5. Tổng giá trị ──
    row = merge_section_header(ws, row, "5.  TỔNG GIÁ TRỊ HỢP ĐỒNG")
    ct = data.get("commercial_terms") or {}
    c_label = ws.cell(row, 1, value="Tổng giá trị")
    c_val   = ws.cell(row, 2, value=ct.get("tong_gia_tri_hop_dong") or "Không xác định trong HĐ")
    style_label(c_label)
    c_val.font = _font(bold=True, size=11, color=C_DARK_BLUE)
    c_val.alignment = _align(h="left", v="center")
    c_val.border = _border()
    ws.row_dimensions[row].height = 22
    row += 2

    return ws


# ─── Sheet 2: Commercial Terms ───

def write_commercial_sheet(ws, data: dict):
    ws.title = "💰 Commercial Terms"
    ct = data.get("commercial_terms") or {}

    # Độ rộng cột (6 cột cho bảng phí)
    for col, w in zip("ABCDEF", [22, 28, 18, 20, 20, 30]):
        ws.column_dimensions[col].width = w

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    style_header_main(ws.cell(row, 1), "COMMERCIAL TERMS — PHÍ & ĐIỀU KIỆN THƯƠNG MẠI")
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Bảng phí dịch vụ ──
    ws.merge_cells(f"A{row}:F{row}")
    style_section(ws.cell(row, 1), "PHÍ DỊCH VỤ (CHI TIẾT THEO KÊNH TT / NGUỒN TIỀN)")
    ws.row_dimensions[row].height = 22; row += 1

    headers = ["Loại phí", "Kênh TT", "Nguồn tiền", "Mức phí", "Điều kiện áp dụng", "Ghi chú"]
    for i, h in enumerate(headers, 1):
        style_table_header(ws.cell(row, i), h)
    ws.row_dimensions[row].height = 22; row += 1

    fees = ct.get("phi_dich_vu") or []
    if fees:
        for idx, f in enumerate(fees):
            alt = idx % 2 == 1
            for col, key in enumerate(
                ["loai_phi","kenh_thanh_toan","nguon_tien","muc_phi","dieu_kien","ghi_chu"], 1
            ):
                style_table_cell(ws.cell(row, col), f.get(key), alt=alt)
            ws.row_dimensions[row].height = 35; row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row, 1, value="⚠️  Không tìm thấy thông tin phí dịch vụ trong hợp đồng")
        c.font = _font(bold=True, color="C00000", size=10)
        c.fill = _fill(C_RED_LIGHT); c.border = _red_border()
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── Ngân sách KM ──
    km = ct.get("ngan_sach_khuyen_mai") or {}
    if km.get("tong_ngan_sach") or km.get("the_le_dieu_kien"):
        ws.merge_cells(f"A{row}:F{row}")
        style_section(ws.cell(row, 1), "NGÂN SÁCH & THỂ LỆ KHUYẾN MẠI")
        ws.row_dimensions[row].height = 22; row += 1
        ws.merge_cells(f"B{row}:F{row}")
        ws.cell(row, 1, value="Tổng ngân sách").font = _font(bold=True); ws.cell(row,1).border = _border()
        ws.cell(row, 2, value=km.get("tong_ngan_sach")).border = _border()
        ws.row_dimensions[row].height = 18; row += 1
        ws.merge_cells(f"B{row}:F{row}")
        lc = ws.cell(row, 1, value="Thể lệ / Điều kiện")
        lc.font = _font(bold=True); lc.border = _border()
        vc = ws.cell(row, 2, value=km.get("the_le_dieu_kien"))
        vc.alignment = _align(); vc.border = _border()
        ws.row_dimensions[row].height = 50; row += 1
        row += 1

    # ── Lãi & Phạt ──
    lp = ct.get("lai_va_phat") or {}
    ws.merge_cells(f"A{row}:F{row}")
    style_section(ws.cell(row, 1), "LÃI TRẢ CHẬM & PHẠT VI PHẠM")
    ws.row_dimensions[row].height = 22; row += 1
    for label, key in [
        ("Lãi trả chậm", "lai_tra_cham"),
        ("Phạt vi phạm", "phat_vi_pham"),
        ("Bồi thường thiệt hại", "boi_thuong_thiet_hai"),
    ]:
        lc = ws.cell(row, 1, value=label); lc.font = _font(bold=True); lc.border = _border()
        ws.merge_cells(f"B{row}:F{row}")
        vc = ws.cell(row, 2, value=lp.get(key) or "—")
        vc.alignment = _align(); vc.border = _border()
        ws.row_dimensions[row].height = 35; row += 1
    row += 1

    return ws


# ─── Sheet 3: Payment & Reconciliation ───

def write_payment_sheet(ws, data: dict):
    ws.title = "🏦 Payment & Recon"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 72
    ct = data.get("commercial_terms") or {}
    pt = ct.get("payment_term") or {}
    rt = ct.get("reconciliation_term") or {}

    row = 1
    ws.merge_cells(f"A{row}:B{row}")
    style_header_main(ws.cell(row, 1), "PAYMENT TERM & RECONCILIATION TERM")
    ws.row_dimensions[row].height = 28; row += 2

    # ── Payment Term ──
    row = merge_section_header(ws, row, "PAYMENT TERM — ĐIỀU KHOẢN THANH TOÁN")
    for label, key, can_alert in [
        ("Cơ chế thanh toán", "co_che_thanh_toan", True),
        ("Tạm ứng / TT trước", "tam_ung_thanh_toan_truoc", False),
        ("Công nợ thanh toán", "cong_no_thanh_toan", True),
    ]:
        val = pt.get(key)
        write_section_row(ws, row, label, val, alert=(can_alert and not val))
        row += 1

    # Hồ sơ TT
    ho_so = pt.get("ho_so_thanh_toan") or []
    write_section_row(ws, row, "Hồ sơ thanh toán",
        ", ".join(ho_so) if ho_so else None,
        alert=not ho_so)
    row += 1

    alert_msg = pt.get("alert_ho_so")
    if alert_msg:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, value=f"⚠️  {alert_msg}")
        c.font = _font(bold=True, color=C_WHITE, size=10)
        c.fill = _fill(C_RED_HEAD)
        c.alignment = _align(h="left", v="center")
        c.border = _border()
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── Reconciliation Term ──
    row = merge_section_header(ws, row, "RECONCILIATION TERM — ĐIỀU KHOẢN ĐỐI SOÁT")
    for label, key in [
        ("Bắt đầu đối soát", "thoi_gian_bat_dau_doi_soat"),
        ("Gửi đối soát", "thoi_gian_gui_doi_soat"),
        ("Thời hạn phản hồi", "thoi_gian_phan_hoi"),
        ("Xử lý chênh lệch", "xu_ly_chenh_lech"),
        ("Zalopay xuất hóa đơn", "zalopay_xuat_hoa_don"),
    ]:
        val = rt.get(key)
        write_section_row(ws, row, label, val, alert=not val)
        row += 1

    return ws


# ─── Sheet 4: Email vs HĐ ───

def write_comparison_sheet(ws, comparison: dict):
    ws.title = "📧 Email vs HĐ"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 32

    row = 1
    ws.merge_cells(f"A{row}:E{row}")
    style_header_main(ws.cell(row, 1), "SO SÁNH: EMAIL ALIGNMENT vs HỢP ĐỒNG")
    ws.row_dimensions[row].height = 28; row += 1

    # Tổng kết
    summary = comparison.get("tong_ket", "")
    if summary:
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, value=f"Nhận xét tổng quan: {summary}")
        c.font = _font(italic=True, size=10)
        c.fill = _fill(C_YELLOW)
        c.alignment = _align(h="left", v="center")
        ws.row_dimensions[row].height = 40; row += 1
    row += 1

    # Bảng so sánh
    for col, (h, w) in enumerate(zip(
        ["Điểm so sánh", "Trong email", "Trong hợp đồng", "Trạng thái", "Ghi chú"],
        [28, 40, 40, 16, 32]
    ), 1):
        style_table_header(ws.cell(row, col), h)
    ws.row_dimensions[row].height = 22; row += 1

    STATUS_COLORS = {
        "KHỚP": C_GREEN_LIGHT,
        "KHÁC BIỆT": C_RED_LIGHT,
        "CHỈ TRONG EMAIL": C_YELLOW,
        "CHỈ TRONG HĐ": C_VERY_LIGHT,
    }

    for idx, item in enumerate(comparison.get("ket_qua_so_sanh") or []):
        status = item.get("trang_thai", "")
        bg = STATUS_COLORS.get(status, C_WHITE)
        alt = idx % 2 == 1
        vals = [
            item.get("diem_so_sanh"),
            item.get("trong_email"),
            item.get("trong_hop_dong"),
            status,
            item.get("ghi_chu"),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, value=val or "")
            c.font = _font(size=9,
                           bold=(col == 4),
                           color="C00000" if status == "KHÁC BIỆT" else "000000")
            c.fill = _fill(bg if col != 1 else (C_GRAY if alt else C_WHITE))
            c.alignment = _align()
            c.border = _border()
        ws.row_dimensions[row].height = 38; row += 1

    return ws


# ─── Xuất Excel tổng hợp ───

def export_excel(data: dict, comparison: dict | None,
                 output_path: str, source_files: list, email_files: list):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    write_summary_sheet(ws1, data, source_files)

    ws2 = wb.create_sheet()
    write_commercial_sheet(ws2, data)

    ws3 = wb.create_sheet()
    write_payment_sheet(ws3, data)

    if comparison:
        ws4 = wb.create_sheet()
        write_comparison_sheet(ws4, comparison)

    # Freeze panes
    for ws in wb.worksheets:
        ws.freeze_panes = "A4"

    wb.save(output_path)
    print(f"  ✓ Đã lưu: {output_path}")


# ═══════════════════════════════════════════════
# PHẦN 4: MAIN
# ═══════════════════════════════════════════════

def get_api_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        key = input("Nhập ANTHROPIC_API_KEY: ").strip()
    if not key:
        sys.exit("Cần ANTHROPIC_API_KEY để chạy agent.")
    return key


def main():
    parser = argparse.ArgumentParser(
        description="Contract Review Agent — Zalopay FP&A",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("--contracts", "-c", nargs="+", default=[],
        help="File hợp đồng .docx (HĐ chính + phụ lục)")
    parser.add_argument("--emails", "-e", nargs="+", default=[],
        help="File email .eml để so sánh với HĐ")
    parser.add_argument("--folder", "-f", default=None,
        help="Thư mục chứa tất cả file .docx và .eml")
    parser.add_argument("--output", "-o", default=None,
        help="Đường dẫn file Excel output")
    args = parser.parse_args()

    if not args.contracts and not args.folder:
        parser.print_help()
        sys.exit(0)

    print(f"\n{'═'*55}")
    print("  CONTRACT REVIEW AGENT — ZALOPAY FP&A")
    print(f"{'═'*55}\n")
    print("📂 Đọc file:\n")

    contract_texts, email_texts = collect_files(
        args.contracts, args.emails, args.folder
    )
    if not contract_texts:
        sys.exit("Không đọc được nội dung hợp đồng.")

    api_key = get_api_key()

    print("\n🤖 Phân tích:\n")
    data = extract_contract_info(contract_texts, api_key)

    comparison = None
    if email_texts:
        comparison = compare_email_contract(email_texts, data, api_key)

    # Missing fields alert
    missing = data.get("truong_con_thieu") or []
    if missing:
        print(f"\n⚠️  Thông tin còn thiếu:\n   " + "\n   ".join(f"• {m}" for m in missing))

    # Output path
    if args.output:
        output_path = args.output
    else:
        first = Path(args.contracts[0] if args.contracts else list(contract_texts.keys())[0])
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(first.parent / f"ContractReview_{ts}.xlsx")

    print("\n📊 Xuất Excel:\n")
    export_excel(
        data, comparison, output_path,
        source_files=args.contracts or list(contract_texts.keys()),
        email_files=args.emails or list(email_texts.keys()),
    )

    print(f"\n✅ Hoàn thành!\n   📁 File: {output_path}")
    if comparison:
        n_diff = sum(
            1 for x in (comparison.get("ket_qua_so_sanh") or [])
            if x.get("trang_thai") == "KHÁC BIỆT"
        )
        if n_diff:
            print(f"   ⚠️  Phát hiện {n_diff} điểm khác biệt giữa email và HĐ.")
    print()


if __name__ == "__main__":
    main()
