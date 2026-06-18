"""
Contract Review Agent — Zalopay FP&A
=====================================
Trích xuất thông tin hợp đồng Word + so sánh email alignment → Excel.

Sử dụng:
    python agent.py --contracts hop_dong.docx phu_luc.docx
    python agent.py --contracts hop_dong.docx --emails aligned.eml
    python agent.py --folder ./ho_so/
    python agent.py --contracts hop_dong.docx --output review.xlsx

Biến môi trường:
    ANTHROPIC_API_KEY   (bắt buộc)
"""

from __future__ import annotations

import argparse
import email as email_lib
import json
import os
import re
import sys
from datetime import datetime
from email import policy as email_policy
from pathlib import Path

# ── Third-party ──────────────────────────────────────────────────────────────
try:
    from docx import Document
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
except ImportError:
    sys.exit("ERROR: pip install python-docx")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

try:
    from openai import OpenAI
except ImportError:
    sys.exit("ERROR: pip install openai")

try:
    import pdfplumber
    _PDF_AVAILABLE = True
except ImportError:
    _PDF_AVAILABLE = False

try:
    import extract_msg as _extract_msg
    _MSG_AVAILABLE = True
except ImportError:
    _MSG_AVAILABLE = False


# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═════════════════════════════════════════════════════════════════════════════

MODEL          = os.environ.get("LLM_MODEL", "qwen/qwen3-5-27b")
LLM_BASE_URL   = os.environ.get("LLM_BASE_URL", "https://maas-llm-aiplatform-hcm.api.vngcloud.vn/v1")
MAX_CONTRACT   = 50_000    # chars — trimmed to stay under 60s gateway timeout
MAX_EMAIL      = 20_000    # chars
MAX_TOKENS_EXT = 6_000
MAX_TOKENS_CMP = 2_000

# Colour palette
C = {
    "dark_blue":   "1F4E79",
    "mid_blue":    "2E75B6",
    "light_blue":  "D6E4F0",
    "very_light":  "EBF3FB",
    "white":       "FFFFFF",
    "red_head":    "C00000",
    "red_light":   "FDECEA",
    "yellow":      "FFF2CC",
    "green_light": "E2EFDA",
    "gray":        "F2F2F2",
}

# ═════════════════════════════════════════════════════════════════════════════
# PROMPTS
# ═════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """Bạn là chuyên gia phân tích hợp đồng cho bộ phận FP&A của Zalopay.
Quan điểm: Đại diện FP&A của Zalopay (CTCP Zion) — đánh giá rủi ro từ phía Zalopay, không cần cân bằng hai bên.
Quy tắc bắt buộc:
- ĐỌC KỸ TOÀN BỘ văn bản từ đầu đến cuối trước khi trích xuất — không bỏ sót điều khoản nào, kể cả điều khoản ở phụ lục hoặc trang cuối.
- Trả về JSON hợp lệ DUY NHẤT, không có bất kỳ giải thích nào ngoài JSON.
- KHÔNG ASSUME: tuyệt đối không suy diễn, không điền thêm thông tin không có trong văn bản gốc. Chỉ đọc hiểu và trích xuất đúng những gì được viết.
- Tóm tắt ngắn gọn, rõ ràng — không dùng từ ngữ mơ hồ, không viết theo cách gây hiểu nhầm hoặc làm thay đổi ý nghĩa gốc.
- Rephrase concisely nhưng KHÔNG thay đổi ý nghĩa, KHÔNG thêm thông tin ngoài văn bản gốc.
- Nếu không tìm thấy: dùng null. Nếu không rõ ràng: ghi "Cần xác nhận thêm".
- Số tiền/tỷ lệ: ghi đầy đủ đơn vị và ký hiệu (VND, USD, %).
- Điều khoản pháp lý: trích ngắn gọn, đủ ý.
- NHẤT QUÁN: Phân tích cùng một văn bản nhiều lần phải cho kết quả không khác nhau quá 10%. Ưu tiên độ chính xác tuyệt đối — không đoán, không biến thể tuỳ tiện."""

EXTRACTION_PROMPT = """\
Phân tích hợp đồng/phụ lục dưới đây và trả về JSON theo cấu trúc sau:

{{
  "loai_hop_dong": "string từ tiêu đề",

  "cac_ben": [
    {{
      "ten_ben":       "Bên A / Bên B / tên gọi trong HĐ",
      "ten_cong_ty":   "tên đầy đủ",
      "nguoi_dai_dien":"họ tên",
      "chuc_vu":       "chức vụ",
      "dia_chi":       "địa chỉ hoặc null",
      "ma_so_thue":    "MST hoặc null"
    }}
  ],

  "thoi_han_hop_dong": {{
    "ngay_ky":           "DD/MM/YYYY hoặc null",
    "ngay_hieu_luc":     "DD/MM/YYYY hoặc null",
    "thoi_gian_hieu_luc":"VD: 12 tháng kể từ ngày hiệu lực",
    "ngay_het_han":      "DD/MM/YYYY hoặc null",
    "dieu_kien_gia_han": "mô tả gia hạn tự động/thủ công, thời hạn thông báo v.v."
  }},

  "dich_vu_hop_tac": {{
    "mo_ta_chung":      "mô tả ngành nghề và phạm vi hợp tác",
    "doi_tuong_tich_hop":"đối tượng/sản phẩm tích hợp",
    "kenh_thanh_toan":  ["Zalopay App", "Gateway", "VietQR", "..."],
    "nguon_tien":       ["Số dư ví", "Tài khoản ngân hàng", "Thẻ nội địa", "Thẻ quốc tế", "..."]
  }},

  "commercial_terms": {{
    "tong_gia_tri_hop_dong": "số tiền + đơn vị hoặc null",

    "phi_dich_vu": [
      {{
        "loai_phi":       "tên loại phí (dịch vụ / hoàn trả / chia sẻ / thúc đẩy DS / ...)",
        "kenh_thanh_toan":"kênh áp dụng",
        "nguon_tien":     "nguồn tiền áp dụng",
        "muc_phi":        "con số / tỷ lệ cụ thể",
        "dieu_kien":      "điều kiện áp dụng ngắn gọn",
        "ghi_chu":        "thông tin bổ sung hoặc null"
      }}
    ],

    "ngan_sach_khuyen_mai": {{
      "tong_ngan_sach":    "số tiền hoặc null",
      "the_le_dieu_kien":  "mô tả thể lệ và điều kiện áp dụng hoặc null"
    }},

    "lai_va_phat": {{
      "lai_tra_cham":          "mức lãi và cách tính hoặc null",
      "phat_vi_pham":          "mức phạt hoặc null",
      "boi_thuong_thiet_hai":  "quy định bồi thường hoặc null"
    }},

    "payment_term": {{
      "co_che_thanh_toan":        "cấn trừ D+1 / đối tác TT sau X ngày từ HĐ đối soát / ...",
      "tam_ung_thanh_toan_truoc": "số tiền hoặc null",
      "cong_no_thanh_toan":       "X ngày thường / X ngày làm việc",
      "ho_so_thanh_toan":         ["hóa đơn VAT", "biên bản đối soát", "biên bản nghiệm thu", "..."],
      "alert_ho_so":              "CẢNH BÁO nếu thiếu hóa đơn / biên bản đối soát / nghiệm thu, hoặc null"
    }},

    "reconciliation_term": {{
      "thoi_gian_bat_dau_doi_soat":"mô tả",
      "thoi_gian_gui_doi_soat":    "mô tả",
      "thoi_gian_phan_hoi":        "mô tả",
      "xu_ly_chenh_lech":          "quy trình xử lý chênh lệch",
      "zalopay_xuat_hoa_don":      "X ngày kể từ xác nhận đối soát hoặc null"
    }}
  }},

  "truong_con_thieu": ["trường quan trọng không tìm thấy trong HĐ"],

  "danh_gia_rui_ro": [
    {{
      "loai_rui_ro": "Tài chính | Pháp lý | Tuân thủ",
      "muc_do":      "CAO | TRUNG_BINH | THAP",
      "tom_tat":     "Tóm tắt rủi ro cho Zalopay — chính xác ngữ nghĩa, tối đa 2 câu",
      "dieu_khoan":  "Điều/Khoản dẫn chiếu hoặc null"
    }}
  ],

  "_refs": {{
    "loai_hop_dong":                        "Điều/Khoản chứa thông tin này, hoặc null",
    "cac_ben":                              "Điều/Khoản hoặc null",
    "thoi_han_hop_dong":                    "Điều/Khoản hoặc null",
    "dich_vu_hop_tac":                      "Điều/Khoản hoặc null",
    "commercial_terms.tong_gia_tri":        "Điều/Khoản hoặc null",
    "commercial_terms.phi_dich_vu":         "Điều/Khoản hoặc null",
    "commercial_terms.ngan_sach_khuyen_mai":"Điều/Khoản hoặc null",
    "commercial_terms.lai_va_phat":         "Điều/Khoản hoặc null",
    "commercial_terms.payment_term":        "Điều/Khoản hoặc null",
    "commercial_terms.reconciliation_term": "Điều/Khoản hoặc null"
  }}
}}

Đánh giá rủi ro cho Zalopay — chỉ liệt kê điều khoản bất lợi thực sự:
- Tài chính CAO: phạt/lãi/bồi thường nặng không có trần phía Zalopay; cơ chế TT gây chậm thu; phí mơ hồ ảnh hưởng doanh thu
- Pháp lý CAO: bồi thường vô hạn phía Zalopay; giới hạn trách nhiệm một chiều; quyền đơn phương sửa đổi của đối tác; điều khoản chấm dứt bất lợi
- Tuân thủ CAO: nghĩa vụ báo cáo/kiểm toán bất khả thi; không phù hợp quy định NHNN/Bộ TC
- TRUNG_BINH/THAP: điều khoản mơ hồ có thể bất lợi, thiếu cơ chế bảo vệ nhỏ
Sắp xếp CAO→TRUNG_BINH→THAP. Nếu không có rủi ro đáng kể: "danh_gia_rui_ro": [].

=== NỘI DUNG TÀI LIỆU ===
{content}
"""

COMPARISON_PROMPT = """\
So sánh nội dung email alignment với hợp đồng. Tuân thủ nghiêm các quy tắc bên dưới.

QUY TẮC ĐỌC EMAIL — BẮT BUỘC:
1. Đọc TOÀN BỘ chuỗi email từ đầu đến cuối theo thứ tự thời gian (email cũ → mới nhất).
2. Lấy alignment MỚI NHẤT: nếu cùng một điểm được sửa đổi nhiều lần, chỉ giữ phiên bản hai bên đồng thuận gần nhất.
3. BỎ QUA hoàn toàn các điểm mà hai bên đã thống nhất loại bỏ (VD: "remove", "bỏ khoản này", "không áp dụng") — KHÔNG đưa vào ket_qua_so_sanh.
4. Đánh dấu trang_thai = "CHUA_PHAN_HOI" cho mọi điểm: (a) chỉ một bên đề xuất nhưng chưa được bên kia xác nhận, hoặc (b) còn đang tranh luận / chưa chốt.

Trả về JSON:
{{
  "ket_qua_so_sanh": [
    {{
      "diem_so_sanh":  "tên điểm so sánh (VD: Phí dịch vụ Gateway)",
      "trong_email":   "nội dung aligned mới nhất trong email, hoặc null",
      "trong_hop_dong":"nội dung trong HĐ, hoặc null nếu không có",
      "trang_thai":    "KHOP | KHAC_BIET | CHI_EMAIL | CHI_HD | CHUA_PHAN_HOI",
      "ghi_chu":       "giải thích ngắn gọn nếu có sai lệch hoặc chưa đồng thuận"
    }}
  ],
  "tong_ket": "nhận xét tổng quan: mức độ alignment và các điểm cần lưu ý nhất"
}}

Chú thích trang_thai:
- KHOP: nội dung email khớp hoàn toàn với HĐ
- KHAC_BIET: nội dung email khác HĐ (đã aligned nhưng HĐ chưa cập nhật)
- CHI_EMAIL: điểm được aligned trong email nhưng không thấy trong HĐ
- CHI_HD: điểm chỉ có trong HĐ, không được đề cập trong email
- CHUA_PHAN_HOI: điểm chưa được đồng thuận rõ ràng từ cả hai bên

=== NỘI DUNG EMAIL ===
{email_content}

=== TÓM TẮT HỢP ĐỒNG ===
{contract_summary}
"""


# ═════════════════════════════════════════════════════════════════════════════
# READERS
# ═════════════════════════════════════════════════════════════════════════════

def read_docx(path: str) -> str:
    """Trả về toàn bộ text từ .docx, bao gồm nội dung bảng."""
    doc = Document(path)
    parts: list[str] = []
    for block in doc.element.body:
        tag = block.tag.split("}")[-1]
        if tag == "p":
            text = "".join(
                node.text or "" for node in block.iter() if node.tag == qn("w:t")
            )
            if text.strip():
                parts.append(text)
        elif tag == "tbl":
            tbl = DocxTable(block, doc)
            for row in tbl.rows:
                row_text = " | ".join(cell.text.strip() for cell in row.cells)
                if row_text.strip(" |"):
                    parts.append(f"[TABLE] {row_text}")
    return "\n".join(parts)


def read_eml(path: str) -> str:
    """Trả về nội dung email (.eml) dạng plain text."""
    with open(path, "rb") as fh:
        msg = email_lib.message_from_bytes(fh.read(), policy=email_policy.default)

    subject = msg.get("Subject", "")
    sender  = msg.get("From", "")
    to      = msg.get("To", "")
    date    = msg.get("Date", "")
    body    = ""

    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                body += part.get_content() or ""
            elif ct == "text/html" and not body:
                html = part.get_content() or ""
                body += re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html))
    else:
        body = msg.get_content() or ""

    return (
        f"[EMAIL]\nTừ: {sender}\nGửi đến: {to}\nNgày: {date}\n"
        f"Tiêu đề: {subject}\n\nNội dung:\n{body}"
    )


def read_pdf(path: str) -> str:
    """Trả về toàn bộ text từ .pdf, bao gồm nội dung bảng."""
    if not _PDF_AVAILABLE:
        raise RuntimeError("pdfplumber chưa được cài: pip install pdfplumber")
    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if text.strip():
                parts.append(text)
            for table in page.extract_tables():
                for row in table:
                    row_text = " | ".join(cell or "" for cell in row)
                    if row_text.strip(" |"):
                        parts.append(f"[TABLE] {row_text}")
    return "\n".join(parts)


def read_msg(path: str) -> str:
    """Trả về nội dung email Outlook (.msg) dạng plain text."""
    if not _MSG_AVAILABLE:
        raise RuntimeError("extract-msg chưa được cài: pip install extract-msg")
    m = _extract_msg.openMsg(path)
    subject = m.subject or ""
    sender  = m.sender or ""
    to      = m.to or ""
    date    = str(m.date or "")
    body    = m.body or ""
    if body.lstrip().startswith("<"):
        body = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", body))
    return (
        f"[EMAIL]\nTừ: {sender}\nGửi đến: {to}\nNgày: {date}\n"
        f"Tiêu đề: {subject}\n\nNội dung:\n{body}"
    )


def collect_files(
    contracts: list[str],
    emails: list[str],
    folder: str | None,
) -> tuple[dict[str, str], dict[str, str]]:
    """Đọc tất cả file, trả về (contract_texts, email_texts)."""
    if folder:
        p = Path(folder)
        contracts = contracts + [str(f) for f in sorted(p.glob("*.docx"))] \
                              + [str(f) for f in sorted(p.glob("*.pdf"))]
        emails    = emails    + [str(f) for f in sorted(p.glob("*.eml"))] \
                              + [str(f) for f in sorted(p.glob("*.msg"))]

    contract_texts: dict[str, str] = {}
    email_texts:    dict[str, str] = {}

    for fp in contracts:
        name = Path(fp).name
        ext  = Path(fp).suffix.lower()
        try:
            contract_texts[name] = read_pdf(fp) if ext == ".pdf" else read_docx(fp)
            print(f"  ✓ HĐ/PL : {name}")
        except Exception as exc:
            print(f"  ✗ Lỗi   : {name} — {exc}")

    for fp in emails:
        name = Path(fp).name
        ext  = Path(fp).suffix.lower()
        try:
            email_texts[name] = read_msg(fp) if ext == ".msg" else read_eml(fp)
            print(f"  ✓ Email : {name}")
        except Exception as exc:
            print(f"  ✗ Lỗi   : {name} — {exc}")

    return contract_texts, email_texts


# ═════════════════════════════════════════════════════════════════════════════
# EXTRACTORS
# ═════════════════════════════════════════════════════════════════════════════

def _call_llm(
    prompt: str,
    api_key: str,
    max_tokens: int,
    system_suffix: str = "",
) -> dict:
    """Gọi LLM API (OpenAI-compatible), parse JSON từ response."""
    client = OpenAI(api_key=api_key, base_url=LLM_BASE_URL)
    system = SYSTEM_PROMPT + (f"\n\n{system_suffix}" if system_suffix else "")
    response = client.chat.completions.create(
        model=MODEL,
        max_tokens=max_tokens,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": prompt},
        ],
        extra_body={"chat_template_kwargs": {"enable_thinking": False}},
    )
    raw = response.choices[0].message.content.strip()
    # Strip markdown code fences nếu có
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def _call_llm_text(
    messages: list[dict],
    api_key: str,
    max_tokens: int = 1_000,
) -> str:
    """Gọi LLM API, trả về text thuần (dùng cho chat)."""
    client = OpenAI(api_key=api_key, base_url=LLM_BASE_URL)
    response = client.chat.completions.create(
        model=MODEL,
        max_tokens=max_tokens,
        messages=messages,
        extra_body={"chat_template_kwargs": {"enable_thinking": False}},
    )
    return response.choices[0].message.content.strip()


def extract_contract(
    texts: dict[str, str],
    api_key: str,
    learning_example: dict | None = None,
) -> dict:
    """Trích xuất thông tin có cấu trúc từ hợp đồng.

    learning_example: JSON result từ một lần review thành công trước đó,
    dùng làm few-shot context để cải thiện độ chính xác.
    """
    full_content = ""
    for name, text in texts.items():
        full_content += f"\n\n{'='*60}\nTÀI LIỆU: {name}\n{'='*60}\n{text}"
    prompt = EXTRACTION_PROMPT.format(content=full_content[:MAX_CONTRACT])
    print("  → Trích xuất thông tin HĐ...")

    system_suffix = ""
    if learning_example:
        example_json = json.dumps(learning_example, ensure_ascii=False, indent=2)[:6_000]
        system_suffix = (
            "=== VÍ DỤ TRÍCH XUẤT THÀNH CÔNG (tham khảo cấu trúc và mức độ chi tiết) ===\n"
            + example_json
        )

    return _call_llm(prompt, api_key, MAX_TOKENS_EXT, system_suffix=system_suffix)


def compare_email(
    email_texts: dict[str, str],
    contract_data: dict,
    api_key: str,
) -> dict:
    """So sánh nội dung email alignment với hợp đồng."""
    email_content    = "\n\n".join(email_texts.values())[:MAX_EMAIL]
    contract_summary = json.dumps(contract_data, ensure_ascii=False, indent=2)[:30_000]
    prompt = COMPARISON_PROMPT.format(
        email_content=email_content,
        contract_summary=contract_summary,
    )
    print("  → So sánh email vs HĐ...")
    return _call_llm(prompt, api_key, MAX_TOKENS_CMP)


CUSTOM_QUERY_PROMPT = """\
Dựa trên hợp đồng dưới đây, hãy trả lời TỪNG câu hỏi.
Trả về JSON hợp lệ DUY NHẤT:

{{
  "answers": [
    {{
      "question": "câu hỏi gốc",
      "answer":   "câu trả lời ngắn gọn, dựa ĐÚNG theo văn bản HĐ",
      "ref":      "Điều/Khoản chứa thông tin, hoặc null"
    }}
  ]
}}

Nếu không tìm thấy thông tin: answer = "Không tìm thấy trong hợp đồng."

=== CÂU HỎI ===
{questions}

=== NỘI DUNG HỢP ĐỒNG ===
{content}
"""


CHAT_SYSTEM_PROMPT = """Bạn là trợ lý phân tích hợp đồng cho bộ phận FP&A của Zalopay.
Bạn đang hỗ trợ người dùng tra cứu và giải đáp thắc mắc về hợp đồng đã được phân tích.
Quy tắc bắt buộc:
- KHÔNG ASSUME: chỉ trả lời dựa trên nội dung hợp đồng được cung cấp. Tuyệt đối không suy diễn.
- Nếu thông tin không có trong hợp đồng: trả lời rõ ràng "Không tìm thấy thông tin này trong hợp đồng."
- Tóm tắt ngắn gọn, rõ ràng — không dùng từ ngữ gây hiểu nhầm.
- Dẫn chiếu Điều/Khoản cụ thể khi có thể.
- Trả lời bằng ngôn ngữ của câu hỏi (thường là tiếng Việt).

Định dạng phản hồi (QUAN TRỌNG — luôn tuân thủ):
- Dùng Markdown: **in đậm** cho số liệu/điều khoản quan trọng, - cho danh sách, bảng dạng | Col1 | Col2 | cho dữ liệu dạng bảng.
- Khi câu trả lời liên quan đến một mục cụ thể trong kết quả phân tích, thêm dòng cuối cùng (dòng riêng biệt):
  [SECTION:tên_mục] — với tên_mục là một trong: summary, commercial, payment, risk, comparison
  Ví dụ: câu hỏi về phí → thêm [SECTION:commercial]; về rủi ro → [SECTION:risk]; về thanh toán/đối soát → [SECTION:payment]
- Giữ câu trả lời súc tích, tối đa 400 từ."""


def chat_with_contract(
    message: str,
    contract_data: dict,
    chat_history: list[dict],
    api_key: str,
) -> dict:
    """Trả lời câu hỏi người dùng dựa trên dữ liệu hợp đồng đã trích xuất.
    Trả về dict {reply: str, highlight: str|None}.
    """
    contract_summary = json.dumps(contract_data, ensure_ascii=False, indent=2)[:30_000]
    system = (
        CHAT_SYSTEM_PROMPT
        + f"\n\n=== TÓM TẮT HỢP ĐỒNG ĐÃ PHÂN TÍCH ===\n{contract_summary}"
    )
    messages: list[dict] = [{"role": "system", "content": system}]
    for turn in (chat_history or [])[-8:]:
        messages.append({"role": turn["role"], "content": turn["content"]})
    messages.append({"role": "user", "content": message})
    raw = _call_llm_text(messages, api_key, max_tokens=1_500)

    # Parse [SECTION:xxx] tag
    import re as _re
    highlight = None
    m = _re.search(r'\[SECTION:([\w]+)\]\s*$', raw.strip())
    if m:
        highlight = m.group(1)
        raw = raw[:m.start()].rstrip()

    return {"reply": raw, "highlight": highlight}


def answer_custom_queries(
    texts: dict[str, str],
    questions: list[str],
    api_key: str,
) -> list[dict]:
    """Trả lời câu hỏi tuỳ chọn từ người dùng dựa trên nội dung hợp đồng."""
    if not questions:
        return []
    content = "\n\n".join(
        f"=== {fname} ===\n{text[:MAX_CONTRACT]}"
        for fname, text in texts.items()
    )
    prompt = CUSTOM_QUERY_PROMPT.format(
        questions="\n".join(f"{i+1}. {q}" for i, q in enumerate(questions)),
        content=content,
    )
    print("  → Trả lời câu hỏi tùy chọn...")
    result = _call_llm(prompt, api_key, 2_000)
    return result.get("answers", [])


GROUP_CONTRACTS_PROMPT = """\
Phân loại danh sách tên file sau thành các nhóm hợp đồng logic.
Mỗi nhóm gồm 1 hợp đồng chính và các phụ lục/addendum liên quan.

Quy tắc:
- File phụ lục thường có trong tên: "phu_luc", "phuluc", "PL", "pl_", "_pl", "annex", "addendum", "amendment", "supplement", "phu luc", "PHLC"
- Nhóm phụ lục với HĐ chính dựa trên: số HĐ giống nhau, tên công ty giống nhau, prefix/suffix tên file tương đồng
- Nếu không xác định được quan hệ → mỗi file là 1 nhóm riêng
- Ưu tiên nhóm nhiều file lại khi có dấu hiệu rõ ràng

Danh sách file: {filenames}

Trả về JSON hợp lệ DUY NHẤT:
{{
  "groups": [
    {{
      "group_name": "tên nhóm ngắn gọn (ví dụ: HĐ ABC Corp #123)",
      "main_file": "tên file hợp đồng chính",
      "files": ["file1.docx", "file2.docx"]
    }}
  ]
}}
"""

EMAIL_TARGET_PROMPT = """\
Đọc nội dung email sau. Xác định email này đang thảo luận/đề cập đến nhóm hợp đồng nào trong danh sách.
Trả về JSON hợp lệ DUY NHẤT:
{{
  "matched_group_names": ["tên nhóm 1", ...],
  "confidence": "HIGH | MEDIUM | LOW",
  "reason": "lý do ngắn gọn (1 câu)"
}}
Nếu không xác định được → trả matched_group_names: [] (mảng rỗng).

=== NỘI DUNG EMAIL ===
{email_content}

=== DANH SÁCH NHÓM HỢP ĐỒNG ===
{groups_info}
"""


def group_contract_files(filenames: list[str], api_key: str) -> list[dict]:
    """Nhóm các file hợp đồng/phụ lục thành các nhóm logic (1 HĐ + phụ lục của nó).
    Trả về list[{group_name, main_file, files}].
    """
    if len(filenames) <= 1:
        name = filenames[0] if filenames else "Hợp đồng"
        return [{"group_name": name, "main_file": name, "files": filenames}]

    prompt = GROUP_CONTRACTS_PROMPT.format(
        filenames=json.dumps(filenames, ensure_ascii=False)
    )
    print("  → Nhóm các file hợp đồng/phụ lục...")
    try:
        result = _call_llm(prompt, api_key, 1_000)
        groups = result.get("groups") or []
        if not groups:
            raise ValueError("empty groups")
        # Validate: mỗi file phải xuất hiện trong ít nhất 1 group
        all_in_groups = {f for g in groups for f in g.get("files", [])}
        for f in filenames:
            if f not in all_in_groups:
                groups.append({"group_name": f, "main_file": f, "files": [f]})
        return groups
    except Exception as exc:
        print(f"  ⚠️  Grouping fallback: {exc}")
        return [{"group_name": f, "main_file": f, "files": [f]} for f in filenames]


def identify_email_target(
    email_texts: dict[str, str],
    groups: list[dict],
    api_key: str,
) -> list[str]:
    """Xác định email đang nói về nhóm hợp đồng nào.
    Trả về list group_names được match. Nếu không xác định → trả toàn bộ group names.
    """
    if len(groups) <= 1:
        return [g["group_name"] for g in groups]

    email_content = "\n\n".join(email_texts.values())[:10_000]
    groups_info = "\n".join(
        f"- {g['group_name']}: {', '.join(g['files'])}" for g in groups
    )
    prompt = EMAIL_TARGET_PROMPT.format(
        email_content=email_content,
        groups_info=groups_info,
    )
    print("  → Xác định hợp đồng liên quan đến email...")
    try:
        result = _call_llm(prompt, api_key, 500)
        matched = result.get("matched_group_names") or []
        if matched:
            return matched
    except Exception:
        pass
    return [g["group_name"] for g in groups]


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border(color: str = "B0C4D8") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="top", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header(cell, text: str, bg=C["dark_blue"], size=13):
    cell.value = text
    cell.font  = _font(bold=True, size=size, color=C["white"])
    cell.fill  = _fill(bg)
    cell.alignment = _align(h="center", v="center")

def _section(cell, text: str, bg=C["mid_blue"]):
    cell.value = text
    cell.font  = _font(bold=True, size=10, color=C["white"])
    cell.fill  = _fill(bg)
    cell.alignment = _align(h="left", v="center")
    cell.border = _thin_border()

def _sub_section(cell, text: str):
    cell.value = text
    cell.font  = _font(bold=True, size=10, color=C["dark_blue"])
    cell.fill  = _fill(C["light_blue"])
    cell.alignment = _align(h="left", v="center")
    cell.border = _thin_border()

def _label(cell, text: str):
    cell.value = text
    cell.font  = _font(bold=True, size=10)
    cell.fill  = _fill(C["very_light"])
    cell.alignment = _align()
    cell.border = _thin_border()

def _value(cell, text, alert=False):
    cell.value = text or ""
    cell.font  = _font(size=10, color=C["red_head"] if alert else "000000")
    cell.fill  = _fill(C["red_light"] if alert else C["white"])
    cell.alignment = _align()
    cell.border = _thin_border("C00000" if alert else "B0C4D8")

def _tbl_header(cell, text: str):
    cell.value = text
    cell.font  = _font(bold=True, size=9, color=C["white"])
    cell.fill  = _fill(C["mid_blue"])
    cell.alignment = _align(h="center", v="center")
    cell.border = _thin_border()

def _tbl_cell(cell, text, alt=False, red=False):
    cell.value = text or ""
    cell.font  = _font(size=9, color=C["red_head"] if red else "000000")
    cell.fill  = _fill(C["red_light"] if red else (C["gray"] if alt else C["white"]))
    cell.alignment = _align()
    cell.border = _thin_border()


def _merge_section(ws, row: int, text: str, n_cols=2, bg=C["mid_blue"]) -> int:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _section(ws.cell(row, 1), text, bg=bg)
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_row(ws, row: int, label: str, value, alert=False, height: int | None = None):
    _label(ws.cell(row, 1), label)
    _value(ws.cell(row, 2), value, alert=alert)
    ws.row_dimensions[row].height = (
        height if height
        else (45 if value and len(str(value)) > 120 else 20)
    )


# ═════════════════════════════════════════════════════════════════════════════
# WRITERS
# ═════════════════════════════════════════════════════════════════════════════

def write_contract_sheet(ws, data: dict, source_files: list[str]) -> None:
    """Ghi toàn bộ thông tin hợp đồng vào 1 sheet duy nhất (6 cột)."""
    NCOLS = 6
    LAST  = "F"
    for col, w in zip("ABCDEF", [30, 28, 20, 20, 30, 28]):
        ws.column_dimensions[col].width = w

    def wr(row, label, value, alert=False, height=None):
        _label(ws.cell(row, 1), label)
        ws.merge_cells(f"B{row}:{LAST}{row}")
        _value(ws.cell(row, 2), value, alert=alert)
        ws.row_dimensions[row].height = (
            height if height else (45 if value and len(str(value)) > 120 else 20)
        )

    def sec(row, text, bg=C["mid_blue"]):
        ws.merge_cells(f"A{row}:{LAST}{row}")
        _section(ws.cell(row, 1), text, bg=bg)
        ws.row_dimensions[row].height = 22
        return row + 1

    row = 1

    # ── Tiêu đề ──
    ws.merge_cells(f"A{row}:{LAST}{row}")
    _header(ws.cell(row, 1), "BẢNG TRÍCH XUẤT THÔNG TIN HỢP ĐỒNG — ZALOPAY FP&A")
    ws.row_dimensions[row].height = 32; row += 1

    ws.merge_cells(f"A{row}:{LAST}{row}")
    c = ws.cell(row, 1, value=(
        f"📅 Ngày xử lý: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
        f"📄 Nguồn: {', '.join(Path(f).name for f in source_files)}"
    ))
    c.font = _font(italic=True, size=9, color="666666")
    c.alignment = _align(h="left", v="center", wrap=False)
    ws.row_dimensions[row].height = 16; row += 1

    missing = data.get("truong_con_thieu") or []
    if missing:
        ws.merge_cells(f"A{row}:{LAST}{row}")
        c = ws.cell(row, 1, value=f"⚠️  THÔNG TIN CÒN THIẾU: {' | '.join(missing)}")
        c.font  = _font(bold=True, size=10, color=C["white"])
        c.fill  = _fill(C["red_head"])
        c.alignment = _align(h="left", v="center")
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── 1. Loại HĐ ──
    row = sec(row, "1.  LOẠI HỢP ĐỒNG")
    wr(row, "Loại hợp đồng", data.get("loai_hop_dong"), alert=not data.get("loai_hop_dong"))
    row += 2

    # ── 2. Các bên ──
    row = sec(row, "2.  CÁC BÊN THAM GIA")
    for party in data.get("cac_ben") or []:
        _label(ws.cell(row, 1), f"Bên: {party.get('ten_ben', '')}")
        ws.merge_cells(f"B{row}:{LAST}{row}")
        _value(ws.cell(row, 2), party.get("ten_cong_ty"))
        ws.row_dimensions[row].height = 18; row += 1
        wr(row, "  Người đại diện",
           f"{party.get('nguoi_dai_dien','')} — {party.get('chuc_vu','')}"); row += 1
        if party.get("ma_so_thue"):
            wr(row, "  MST", party.get("ma_so_thue")); row += 1
    row += 1

    # ── 3. Thời hạn ──
    row = sec(row, "3.  THỜI HẠN HỢP ĐỒNG")
    t = data.get("thoi_han_hop_dong") or {}
    for lbl, key in [
        ("Ngày ký",            "ngay_ky"),
        ("Ngày hiệu lực",      "ngay_hieu_luc"),
        ("Thời gian hiệu lực", "thoi_gian_hieu_luc"),
        ("Ngày hết hạn",       "ngay_het_han"),
        ("Điều kiện gia hạn",  "dieu_kien_gia_han"),
    ]:
        wr(row, lbl, t.get(key), alert=not t.get(key)); row += 1
    row += 1

    # ── 4. Dịch vụ ──
    row = sec(row, "4.  DỊCH VỤ HỢP TÁC")
    dv = data.get("dich_vu_hop_tac") or {}
    wr(row, "Mô tả chung",        dv.get("mo_ta_chung"),        height=55); row += 1
    wr(row, "Đối tượng tích hợp", dv.get("doi_tuong_tich_hop"));            row += 1
    wr(row, "Kênh thanh toán",
       ", ".join(dv.get("kenh_thanh_toan") or []) or None,
       alert=not dv.get("kenh_thanh_toan")); row += 1
    wr(row, "Nguồn tiền",
       ", ".join(dv.get("nguon_tien") or []) or None,
       alert=not dv.get("nguon_tien")); row += 1
    row += 1

    # ── 5. Tổng giá trị ──
    row = sec(row, "5.  TỔNG GIÁ TRỊ HỢP ĐỒNG")
    ct = data.get("commercial_terms") or {}
    _label(ws.cell(row, 1), "Tổng giá trị")
    ws.merge_cells(f"B{row}:{LAST}{row}")
    c = ws.cell(row, 2, value=ct.get("tong_gia_tri_hop_dong") or "Không xác định trong HĐ")
    c.font = _font(bold=True, size=11, color=C["dark_blue"])
    c.alignment = _align(h="left", v="center"); c.border = _thin_border()
    ws.row_dimensions[row].height = 22; row += 2

    # ── 6. Phí dịch vụ ──
    row = sec(row, "6.  PHÍ DỊCH VỤ (THEO KÊNH TT / NGUỒN TIỀN)")
    for col, hdr in enumerate(
        ["Loại phí","Kênh TT","Nguồn tiền","Mức phí","Điều kiện áp dụng","Ghi chú"], 1
    ):
        _tbl_header(ws.cell(row, col), hdr)
    ws.row_dimensions[row].height = 22; row += 1
    fees = ct.get("phi_dich_vu") or []
    if fees:
        for idx, fee in enumerate(fees):
            alt = idx % 2 == 1
            for col, key in enumerate(
                ["loai_phi","kenh_thanh_toan","nguon_tien","muc_phi","dieu_kien","ghi_chu"], 1
            ):
                _tbl_cell(ws.cell(row, col), fee.get(key), alt=alt)
            ws.row_dimensions[row].height = 35; row += 1
    else:
        ws.merge_cells(f"A{row}:{LAST}{row}")
        c = ws.cell(row, 1, value="⚠️  Không tìm thấy thông tin phí dịch vụ trong hợp đồng")
        c.font = _font(bold=True, color=C["red_head"], size=10)
        c.fill = _fill(C["red_light"]); c.border = _thin_border("C00000")
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── 7. Ngân sách KM ──
    km = ct.get("ngan_sach_khuyen_mai") or {}
    if km.get("tong_ngan_sach") or km.get("the_le_dieu_kien"):
        row = sec(row, "7.  NGÂN SÁCH & THỂ LỆ KHUYẾN MẠI")
        for lbl, key in [("Tổng ngân sách","tong_ngan_sach"),
                         ("Thể lệ / Điều kiện","the_le_dieu_kien")]:
            _label(ws.cell(row, 1), lbl)
            ws.merge_cells(f"B{row}:{LAST}{row}")
            c = ws.cell(row, 2, value=km.get(key) or "")
            c.alignment = _align(); c.border = _thin_border()
            ws.row_dimensions[row].height = 50 if key == "the_le_dieu_kien" else 20; row += 1
        row += 1

    # ── 8. Lãi & Phạt ──
    lp = ct.get("lai_va_phat") or {}
    row = sec(row, "8.  LÃI TRẢ CHẬM & PHẠT VI PHẠM")
    for lbl, key in [("Lãi trả chậm","lai_tra_cham"),
                     ("Phạt vi phạm","phat_vi_pham"),
                     ("Bồi thường thiệt hại","boi_thuong_thiet_hai")]:
        _label(ws.cell(row, 1), lbl)
        ws.merge_cells(f"B{row}:{LAST}{row}")
        c = ws.cell(row, 2, value=lp.get(key) or "—")
        c.alignment = _align(); c.border = _thin_border()
        ws.row_dimensions[row].height = 35; row += 1
    row += 1

    # ── 9. Payment Term ──
    pt = ct.get("payment_term") or {}
    row = sec(row, "9.  PAYMENT TERM — ĐIỀU KHOẢN THANH TOÁN")
    for lbl, key, req in [
        ("Cơ chế thanh toán",  "co_che_thanh_toan",        True),
        ("Tạm ứng / TT trước", "tam_ung_thanh_toan_truoc", False),
        ("Công nợ thanh toán", "cong_no_thanh_toan",       True),
    ]:
        wr(row, lbl, pt.get(key), alert=(req and not pt.get(key))); row += 1
    ho_so = pt.get("ho_so_thanh_toan") or []
    wr(row, "Hồ sơ thanh toán", ", ".join(ho_so) if ho_so else None, alert=not ho_so); row += 1
    alert_msg = pt.get("alert_ho_so")
    if alert_msg:
        ws.merge_cells(f"A{row}:{LAST}{row}")
        c = ws.cell(row, 1, value=f"⚠️  {alert_msg}")
        c.font = _font(bold=True, color=C["white"], size=10)
        c.fill = _fill(C["red_head"])
        c.alignment = _align(h="left", v="center"); c.border = _thin_border()
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── 10. Reconciliation ──
    rt = ct.get("reconciliation_term") or {}
    row = sec(row, "10. RECONCILIATION TERM — ĐIỀU KHOẢN ĐỐI SOÁT")
    for lbl, key in [
        ("Bắt đầu đối soát",    "thoi_gian_bat_dau_doi_soat"),
        ("Gửi đối soát",         "thoi_gian_gui_doi_soat"),
        ("Thời hạn phản hồi",    "thoi_gian_phan_hoi"),
        ("Xử lý chênh lệch",     "xu_ly_chenh_lech"),
        ("Zalopay xuất hóa đơn", "zalopay_xuat_hoa_don"),
    ]:
        wr(row, lbl, rt.get(key), alert=not rt.get(key)); row += 1


def write_summary_sheet(ws, data: dict, source_files: list[str]):
    ws.title = "📋 Tóm tắt HĐ"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 72
    row = 1

    # ── Tiêu đề ──
    ws.merge_cells(f"A{row}:B{row}")
    _header(ws.cell(row, 1), "BẢNG TRÍCH XUẤT THÔNG TIN HỢP ĐỒNG — ZALOPAY FP&A")
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1,
        value=(f"📅 Ngày xử lý: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
               f"📄 Nguồn: {', '.join(Path(f).name for f in source_files)}"))
    c.font = _font(italic=True, size=9, color="666666")
    c.alignment = _align(h="left", v="center", wrap=False)
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Alert thiếu thông tin ──
    missing = data.get("truong_con_thieu") or []
    if missing:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, value=f"⚠️  THÔNG TIN CÒN THIẾU: {' | '.join(missing)}")
        c.font  = _font(bold=True, size=10, color=C["white"])
        c.fill  = _fill(C["red_head"])
        c.alignment = _align(h="left", v="center")
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ── 1. Loại HĐ ──
    row = _merge_section(ws, row, "1.  LOẠI HỢP ĐỒNG")
    _write_row(ws, row, "Loại hợp đồng", data.get("loai_hop_dong"),
               alert=not data.get("loai_hop_dong"))
    row += 2

    # ── 2. Các Bên ──
    row = _merge_section(ws, row, "2.  CÁC BÊN THAM GIA")
    for party in data.get("cac_ben") or []:
        _label(ws.cell(row, 1), f"Bên: {party.get('ten_ben', '')}")
        _value(ws.cell(row, 2), party.get("ten_cong_ty"))
        ws.row_dimensions[row].height = 18
        row += 1
        _write_row(ws, row, "  Người đại diện",
                   f"{party.get('nguoi_dai_dien', '')} — {party.get('chuc_vu', '')}")
        row += 1
        if party.get("ma_so_thue"):
            _write_row(ws, row, "  MST", party.get("ma_so_thue"))
            row += 1
    row += 1

    # ── 3. Thời hạn ──
    row = _merge_section(ws, row, "3.  THỜI HẠN HỢP ĐỒNG")
    t = data.get("thoi_han_hop_dong") or {}
    for label, key in [
        ("Ngày ký",            "ngay_ky"),
        ("Ngày hiệu lực",      "ngay_hieu_luc"),
        ("Thời gian hiệu lực", "thoi_gian_hieu_luc"),
        ("Ngày hết hạn",       "ngay_het_han"),
        ("Điều kiện gia hạn",  "dieu_kien_gia_han"),
    ]:
        _write_row(ws, row, label, t.get(key), alert=not t.get(key))
        row += 1
    row += 1

    # ── 4. Dịch vụ hợp tác ──
    row = _merge_section(ws, row, "4.  DỊCH VỤ HỢP TÁC")
    dv = data.get("dich_vu_hop_tac") or {}
    _write_row(ws, row, "Mô tả chung",       dv.get("mo_ta_chung"),       height=55); row += 1
    _write_row(ws, row, "Đối tượng tích hợp",dv.get("doi_tuong_tich_hop")); row += 1
    _write_row(ws, row, "Kênh thanh toán",
               ", ".join(dv.get("kenh_thanh_toan") or []) or None,
               alert=not dv.get("kenh_thanh_toan")); row += 1
    _write_row(ws, row, "Nguồn tiền",
               ", ".join(dv.get("nguon_tien") or []) or None,
               alert=not dv.get("nguon_tien")); row += 1
    row += 1

    # ── 5. Tổng giá trị ──
    row = _merge_section(ws, row, "5.  TỔNG GIÁ TRỊ HỢP ĐỒNG")
    ct  = data.get("commercial_terms") or {}
    _label(ws.cell(row, 1), "Tổng giá trị")
    c = ws.cell(row, 2, value=ct.get("tong_gia_tri_hop_dong") or "Không xác định trong HĐ")
    c.font = _font(bold=True, size=11, color=C["dark_blue"])
    c.alignment = _align(h="left", v="center")
    c.border = _thin_border()
    ws.row_dimensions[row].height = 22
    row += 1


def write_commercial_sheet(ws, data: dict):
    ws.title = "💰 Commercial Terms"
    ct = data.get("commercial_terms") or {}

    col_letters = "ABCDEF"
    col_widths  = [24, 28, 20, 20, 30, 28]
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    _header(ws.cell(row, 1), "COMMERCIAL TERMS — PHÍ & ĐIỀU KIỆN THƯƠNG MẠI")
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Bảng phí dịch vụ ──
    ws.merge_cells(f"A{row}:F{row}")
    _section(ws.cell(row, 1), "PHÍ DỊCH VỤ (CHI TIẾT THEO KÊNH TT / NGUỒN TIỀN)")
    ws.row_dimensions[row].height = 22; row += 1

    for col, header in enumerate(
        ["Loại phí", "Kênh TT", "Nguồn tiền", "Mức phí", "Điều kiện áp dụng", "Ghi chú"], 1
    ):
        _tbl_header(ws.cell(row, col), header)
    ws.row_dimensions[row].height = 22; row += 1

    fees = ct.get("phi_dich_vu") or []
    if fees:
        for idx, fee in enumerate(fees):
            alt = idx % 2 == 1
            for col, key in enumerate(
                ["loai_phi","kenh_thanh_toan","nguon_tien","muc_phi","dieu_kien","ghi_chu"], 1
            ):
                _tbl_cell(ws.cell(row, col), fee.get(key), alt=alt)
            ws.row_dimensions[row].height = 35; row += 1
    else:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row, 1, value="⚠️  Không tìm thấy thông tin phí dịch vụ trong hợp đồng")
        c.font = _font(bold=True, color=C["red_head"], size=10)
        c.fill = _fill(C["red_light"])
        c.border = _thin_border("C00000")
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── Ngân sách KM ──
    km = ct.get("ngan_sach_khuyen_mai") or {}
    if km.get("tong_ngan_sach") or km.get("the_le_dieu_kien"):
        ws.merge_cells(f"A{row}:F{row}")
        _section(ws.cell(row, 1), "NGÂN SÁCH & THỂ LỆ KHUYẾN MẠI")
        ws.row_dimensions[row].height = 22; row += 1
        for label, key in [("Tổng ngân sách","tong_ngan_sach"),("Thể lệ / Điều kiện","the_le_dieu_kien")]:
            _label(ws.cell(row, 1), label)
            ws.merge_cells(f"B{row}:F{row}")
            c = ws.cell(row, 2, value=km.get(key) or "")
            c.alignment = _align(); c.border = _thin_border()
            h = 50 if key == "the_le_dieu_kien" else 20
            ws.row_dimensions[row].height = h; row += 1
        row += 1

    # ── Lãi & Phạt ──
    lp = ct.get("lai_va_phat") or {}
    ws.merge_cells(f"A{row}:F{row}")
    _section(ws.cell(row, 1), "LÃI TRẢ CHẬM & PHẠT VI PHẠM")
    ws.row_dimensions[row].height = 22; row += 1
    for label, key in [
        ("Lãi trả chậm",         "lai_tra_cham"),
        ("Phạt vi phạm",         "phat_vi_pham"),
        ("Bồi thường thiệt hại", "boi_thuong_thiet_hai"),
    ]:
        _label(ws.cell(row, 1), label)
        ws.merge_cells(f"B{row}:F{row}")
        c = ws.cell(row, 2, value=lp.get(key) or "—")
        c.alignment = _align(); c.border = _thin_border()
        ws.row_dimensions[row].height = 35; row += 1


def write_payment_sheet(ws, data: dict):
    ws.title = "🏦 Payment & Recon"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 72
    ct = data.get("commercial_terms") or {}
    pt = ct.get("payment_term")       or {}
    rt = ct.get("reconciliation_term") or {}
    row = 1

    ws.merge_cells(f"A{row}:B{row}")
    _header(ws.cell(row, 1), "PAYMENT TERM & RECONCILIATION TERM")
    ws.row_dimensions[row].height = 28; row += 2

    # ── Payment Term ──
    row = _merge_section(ws, row, "PAYMENT TERM — ĐIỀU KHOẢN THANH TOÁN")
    for label, key, required in [
        ("Cơ chế thanh toán",        "co_che_thanh_toan",        True),
        ("Tạm ứng / TT trước",       "tam_ung_thanh_toan_truoc", False),
        ("Công nợ thanh toán",        "cong_no_thanh_toan",       True),
    ]:
        val = pt.get(key)
        _write_row(ws, row, label, val, alert=(required and not val))
        row += 1

    ho_so = pt.get("ho_so_thanh_toan") or []
    _write_row(ws, row, "Hồ sơ thanh toán",
               ", ".join(ho_so) if ho_so else None, alert=not ho_so)
    row += 1

    alert_msg = pt.get("alert_ho_so")
    if alert_msg:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, value=f"⚠️  {alert_msg}")
        c.font = _font(bold=True, color=C["white"], size=10)
        c.fill = _fill(C["red_head"])
        c.alignment = _align(h="left", v="center")
        c.border = _thin_border()
        ws.row_dimensions[row].height = 22; row += 1
    row += 1

    # ── Reconciliation Term ──
    row = _merge_section(ws, row, "RECONCILIATION TERM — ĐIỀU KHOẢN ĐỐI SOÁT")
    for label, key in [
        ("Bắt đầu đối soát",       "thoi_gian_bat_dau_doi_soat"),
        ("Gửi đối soát",            "thoi_gian_gui_doi_soat"),
        ("Thời hạn phản hồi",       "thoi_gian_phan_hoi"),
        ("Xử lý chênh lệch",        "xu_ly_chenh_lech"),
        ("Zalopay xuất hóa đơn",    "zalopay_xuat_hoa_don"),
    ]:
        _write_row(ws, row, label, rt.get(key), alert=not rt.get(key))
        row += 1


def write_comparison_sheet(ws, comparison: dict):
    ws.title = "📧 Email vs HĐ"
    for col, width in zip("ABCDE", [28, 40, 40, 16, 32]):
        ws.column_dimensions[col].width = width
    row = 1

    ws.merge_cells(f"A{row}:E{row}")
    _header(ws.cell(row, 1), "SO SÁNH: EMAIL ALIGNMENT vs HỢP ĐỒNG")
    ws.row_dimensions[row].height = 28; row += 1

    summary = comparison.get("tong_ket", "")
    if summary:
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, value=f"Nhận xét tổng quan: {summary}")
        c.font = _font(italic=True, size=10)
        c.fill = _fill(C["yellow"])
        c.alignment = _align(h="left", v="center")
        ws.row_dimensions[row].height = 40; row += 1
    row += 1

    for col, header in enumerate(
        ["Điểm so sánh", "Trong email", "Trong hợp đồng", "Trạng thái", "Ghi chú"], 1
    ):
        _tbl_header(ws.cell(row, col), header)
    ws.row_dimensions[row].height = 22; row += 1

    STATUS_BG = {
        "KHỚP":            C["green_light"],
        "KHÁC BIỆT":       C["red_light"],
        "CHỈ TRONG EMAIL": C["yellow"],
        "CHỈ TRONG HĐ":   C["very_light"],
    }

    for idx, item in enumerate(comparison.get("ket_qua_so_sanh") or []):
        status = item.get("trang_thai", "")
        bg = STATUS_BG.get(status, C["white"])
        alt = idx % 2 == 1
        values = [
            item.get("diem_so_sanh"),
            item.get("trong_email"),
            item.get("trong_hop_dong"),
            status,
            item.get("ghi_chu"),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row, col, value=val or "")
            c.font = _font(
                size=9, bold=(col == 4),
                color=C["red_head"] if status == "KHÁC BIỆT" else "000000"
            )
            c.fill = _fill(bg if col != 1 else (C["gray"] if alt else C["white"]))
            c.alignment = _align()
            c.border = _thin_border()
        ws.row_dimensions[row].height = 38; row += 1


def export_excel(
    results: list[dict],
    comparison: dict | None,
    output_path: str,
    source_files: list[str],
):
    """Xuất Excel: 1 sheet mỗi hợp đồng + 1 sheet so sánh email (nếu có)."""
    wb    = openpyxl.Workbook()
    multi = len(results) > 1

    for i, data in enumerate(results):
        fname = source_files[i] if i < len(source_files) else f"hop_dong_{i+1}.docx"
        stem  = Path(fname).stem[:28]

        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = (f"{i+1}. {stem}"[:31] if multi else "📋 Hợp đồng")
        write_contract_sheet(ws, data, [fname])

    if comparison:
        write_comparison_sheet(wb.create_sheet(), comparison)

    for ws in wb.worksheets:
        ws.freeze_panes = "A4"
    wb.save(output_path)
    print(f"  ✓ Lưu : {output_path}")


# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def get_api_key() -> str:
    key = os.environ.get("LLM_API_KEY", "")
    if not key:
        key = input("LLM_API_KEY: ").strip()
    if not key:
        sys.exit("Cần LLM_API_KEY.")
    return key


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Contract Review Agent — Zalopay FP&A",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--contracts", "-c", nargs="+", default=[],
                   metavar="FILE.docx",
                   help="Hợp đồng chính và phụ lục (.docx)")
    p.add_argument("--emails", "-e", nargs="+", default=[],
                   metavar="FILE.eml",
                   help="Email alignment để so sánh (.eml)")
    p.add_argument("--folder", "-f", default=None,
                   metavar="DIR",
                   help="Thư mục chứa toàn bộ .docx và .eml")
    p.add_argument("--output", "-o", default=None,
                   metavar="OUTPUT.xlsx",
                   help="Đường dẫn file Excel output")
    return p


def main():
    args = build_parser().parse_args()

    if not args.contracts and not args.folder:
        build_parser().print_help()
        sys.exit(0)

    print(f"\n{'═' * 55}")
    print("  CONTRACT REVIEW AGENT — ZALOPAY FP&A")
    print(f"{'═' * 55}\n")

    print("📂 Đọc file:\n")
    contract_texts, email_texts = collect_files(args.contracts, args.emails, args.folder)
    if not contract_texts:
        sys.exit("Không đọc được nội dung hợp đồng.")

    api_key = get_api_key()

    print("\n🤖 Phân tích:\n")
    fnames  = list(contract_texts.keys())
    results: list[dict] = []
    for fname, text in contract_texts.items():
        print(f"  → Phân tích: {fname}")
        r = extract_contract({fname: text}, api_key)
        results.append(r)
        missing = r.get("truong_con_thieu") or []
        if missing:
            print(f"     ⚠️  Thiếu: {', '.join(missing)}")

    comparison = None
    if email_texts:
        # So sánh email với hợp đồng đầu tiên (hoặc hợp đồng duy nhất)
        comparison = compare_email(email_texts, results[0], api_key)

    # Xác định output path
    if args.output:
        output_path = args.output
    else:
        first = Path(args.contracts[0]) if args.contracts else Path(fnames[0])
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(first.parent / f"ContractReview_{ts}.xlsx")

    print("\n📊 Xuất Excel:\n")
    export_excel(results, comparison, output_path, fnames)

    print(f"\n✅ Hoàn thành!\n   📁 {output_path}")
    if len(results) > 1:
        print(f"   📑 {len(results)} hợp đồng được phân tích riêng biệt → {len(results)*3} sheets")
    if comparison:
        n_diff = sum(
            1 for x in (comparison.get("ket_qua_so_sanh") or [])
            if x.get("trang_thai") == "KHÁC BIỆT"
        )
        if n_diff:
            print(f"   ⚠️  {n_diff} điểm khác biệt giữa email và HĐ (xem sheet 📧 Email vs HĐ)")
    print()


if __name__ == "__main__":
    main()
